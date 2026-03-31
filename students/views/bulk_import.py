"""
Bulk import students from CSV/Excel with auto-generated credentials.
Preview → validate → skip duplicates → transaction save.
"""
import csv
import io
import re
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth import get_user_model
from django.db import transaction
from django.http import HttpResponse
from accounts.permissions import IsTeacher
from students.models import StudentProfile, ParentProfile, ParentChild, ImportedCredentialRecord
from students.credential_crypto import encrypt_credentials

User = get_user_model()

# Optional: openpyxl for Excel
try:
    import openpyxl
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False


def _normalize_name(name):
    """Normalize name for duplicate detection."""
    return " ".join((name or "").strip().split()).lower()


def _parse_csv(content):
    """Parse CSV content. Returns list of dicts with keys full_name, grade, phone."""
    reader = csv.DictReader(io.StringIO(content))
    fieldnames = reader.fieldnames or []
    name_col = "full_name" if "full_name" in fieldnames else "fullName"
    if name_col not in fieldnames:
        return None, "CSV must have column: full_name or fullName"

    rows = []
    for i, row in enumerate(reader, start=2):
        full_name = (row.get("full_name") or row.get("fullName") or "").strip()
        grade = (row.get("grade") or row.get("class") or "").strip() or None
        phone = (row.get("phone") or "").strip() or None
        rows.append({"row": i, "full_name": full_name, "grade": grade, "phone": phone})
    return rows, None


def _parse_excel(content_bytes):
    """Parse Excel (.xlsx) content. Returns list of dicts or error."""
    if not HAS_EXCEL:
        return None, "Excel support requires openpyxl. Install: pip install openpyxl"
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    except Exception as e:
        return None, f"Invalid Excel file: {e}"
    ws = wb.active
    if not ws:
        return None, "Excel sheet is empty"
    header = [str(c).strip().lower().replace(" ", "_") for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)) or []]
    name_col = "full_name" if "full_name" in header else "fullname"
    if "full_name" not in header and "fullname" not in header:
        wb.close()
        return None, "Excel must have column: full_name or fullName"
    col_idx = header.index("full_name") if "full_name" in header else header.index("fullname")
    grade_idx = header.index("grade") if "grade" in header else (header.index("class") if "class" in header else -1)
    phone_idx = header.index("phone") if "phone" in header else -1

    rows = []
    for i, row_tuple in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_list = list(row_tuple) if row_tuple else []
        full_name = str(row_list[col_idx]).strip() if col_idx < len(row_list) else ""
        grade = str(row_list[grade_idx]).strip() if grade_idx >= 0 and grade_idx < len(row_list) else None
        grade = grade or None
        phone = str(row_list[phone_idx]).strip() if phone_idx >= 0 and phone_idx < len(row_list) else None
        phone = phone or None
        rows.append({"row": i, "full_name": full_name, "grade": grade, "phone": phone})
    wb.close()
    return rows, None


def _validate_row(row):
    """Validate a single row. Returns (valid, error_message)."""
    name = (row.get("full_name") or "").strip()
    if not name:
        return False, "full_name required"
    if len(name) > 255:
        return False, "full_name too long"
    if len(name) < 2:
        return False, "full_name too short"
    phone = row.get("phone")
    if phone and (len(str(phone)) > 20 or not re.match(r"^[\d\s\-\+\(\)\.]*$", str(phone))):
        return False, "invalid phone"
    return True, None


def _get_preview(rows, org):
    """
    Validate rows, detect duplicates (in-file and in-DB).
    Returns: valid_rows, preview_rows (with status for each)
    """
    seen_names = set()
    existing_students = set(
        User.objects.filter(organization=org, role="student")
        .values_list("full_name", flat=True)
    )
    existing_normalized = {_normalize_name(n) for n in existing_students}

    preview_rows = []
    valid_rows = []

    for r in rows:
        full_name = (r.get("full_name") or "").strip()
        normalized = _normalize_name(full_name)
        valid, err = _validate_row(r)

        if not valid:
            preview_rows.append({
                "row": r["row"],
                "fullName": full_name or "(boş)",
                "grade": r.get("grade"),
                "phone": r.get("phone"),
                "status": "invalid",
                "message": err,
            })
            continue

        if normalized in seen_names:
            preview_rows.append({
                "row": r["row"],
                "fullName": full_name,
                "grade": r.get("grade"),
                "phone": r.get("phone"),
                "status": "duplicate_in_file",
                "message": "Bu ad faylda təkrarlanır",
            })
            continue
        seen_names.add(normalized)

        if normalized in existing_normalized:
            preview_rows.append({
                "row": r["row"],
                "fullName": full_name,
                "grade": r.get("grade"),
                "phone": r.get("phone"),
                "status": "duplicate_in_db",
                "message": "Bu şagird artıq mövcuddur",
            })
            continue

        preview_rows.append({
            "row": r["row"],
            "fullName": full_name,
            "grade": r.get("grade"),
            "phone": r.get("phone"),
            "status": "valid",
            "message": None,
        })
        valid_rows.append({"full_name": full_name, "grade": r.get("grade"), "phone": r.get("phone")})

    return valid_rows, preview_rows


def _create_student_with_credentials(org, full_name, grade, phone, created_by=None):
    from students.credentials import generate_credentials, generate_parent_credentials

    creds = generate_credentials(full_name)
    for _ in range(5):
        if User.objects.filter(email=creds["student_email"]).exists():
            creds = generate_credentials(full_name)
            continue
        break
    if User.objects.filter(email=creds["student_email"]).exists():
        return None, None

    student_user = User.objects.create_user(
        email=creds["student_email"],
        password=creds["student_password"],
        full_name=full_name,
        phone=phone,
        role="student",
        is_active=True,
        organization=org,
        must_change_password=True,
    )
    student_profile = StudentProfile.objects.create(
        user=student_user,
        grade=grade,
        balance=0,
    )
    for _ in range(5):
        if User.objects.filter(email=creds["parent_email"]).exists():
            creds["parent_email"], creds["parent_password"] = generate_parent_credentials(full_name)
            continue
        break
    parent_user = User.objects.create_user(
        email=creds["parent_email"],
        password=creds["parent_password"],
        full_name=f"{full_name} — Valideyn",
        role="parent",
        is_active=True,
        organization=org,
        must_change_password=True,
    )
    ParentProfile.objects.create(user=parent_user)
    ParentChild.objects.create(parent=parent_user, student=student_user)

    # Store in credential registry (encrypted)
    try:
        encrypted = encrypt_credentials(creds["student_password"], creds["parent_password"])
        ImportedCredentialRecord.objects.create(
            created_by=created_by,
            source=ImportedCredentialRecord.SOURCE_CSV_IMPORT,
            student=student_user,
            parent=parent_user,
            student_full_name=full_name,
            student_email=creds["student_email"],
            parent_email=creds["parent_email"],
            grade=grade,
            initial_password_encrypted=encrypted,
            password_is_one_time=True,
        )
    except Exception:
        pass  # Do not fail import if registry save fails

    return student_profile, {
        "fullName": full_name,
        "studentEmail": creds["student_email"],
        "studentPassword": creds["student_password"],
        "parentEmail": creds["parent_email"],
        "parentPassword": creds["parent_password"],
    }


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsTeacher])
def bulk_import_preview_view(request):
    """
    POST /api/teacher/bulk-import/preview
    Body: multipart/form-data with file (CSV or .xlsx)
    Returns: preview rows with validation status (valid, invalid, duplicate_in_file, duplicate_in_db)
    """
    if "file" not in request.FILES and "csv" not in request.FILES:
        return Response({"detail": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)
    file = request.FILES.get("file") or request.FILES.get("csv")
    name = (file.name or "").lower()

    if name.endswith(".csv"):
        try:
            content = file.read().decode("utf-8-sig")
        except UnicodeDecodeError:
            return Response({"detail": "Invalid encoding. Use UTF-8."}, status=status.HTTP_400_BAD_REQUEST)
        rows, err = _parse_csv(content)
    elif name.endswith((".xlsx", ".xls")):
        if name.endswith(".xls"):
            return Response({"detail": "Only .xlsx is supported. Save as Excel 2007+."}, status=status.HTTP_400_BAD_REQUEST)
        content_bytes = file.read()
        rows, err = _parse_excel(content_bytes)
    else:
        return Response({"detail": "File must be CSV or Excel (.xlsx)"}, status=status.HTTP_400_BAD_REQUEST)

    if err:
        return Response({"detail": err}, status=status.HTTP_400_BAD_REQUEST)
    if not rows:
        return Response({"detail": "File has no data rows"}, status=status.HTTP_400_BAD_REQUEST)

    org = request.user.organization
    valid_rows, preview_rows = _get_preview(rows, org)

    valid_count = sum(1 for p in preview_rows if p["status"] == "valid")
    invalid_count = sum(1 for p in preview_rows if p["status"] == "invalid")
    dup_file_count = sum(1 for p in preview_rows if p["status"] == "duplicate_in_file")
    dup_db_count = sum(1 for p in preview_rows if p["status"] == "duplicate_in_db")

    return Response({
        "preview": preview_rows,
        "summary": {
            "total": len(preview_rows),
            "valid": valid_count,
            "invalid": invalid_count,
            "duplicateInFile": dup_file_count,
            "duplicateInDb": dup_db_count,
        },
        "validRows": valid_rows,
    }, status=status.HTTP_200_OK)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsTeacher])
def bulk_import_confirm_view(request):
    """
    POST /api/teacher/bulk-import/confirm
    Body: JSON { "rows": [ { "full_name", "grade?", "phone?" } ] }
    Creates students in transaction. Returns created count, errors, credentials.
    """
    rows = request.data.get("rows") or []
    if not isinstance(rows, list):
        return Response({"detail": "rows must be an array"}, status=status.HTTP_400_BAD_REQUEST)

    org = request.user.organization
    created = 0
    errors = []
    credentials_list = []

    with transaction.atomic():
        for i, row in enumerate(rows):
            full_name = (row.get("full_name") or row.get("fullName") or "").strip()
            grade = (row.get("grade") or row.get("class") or "").strip() or None
            phone = (row.get("phone") or "").strip() or None

            valid, err = _validate_row({"full_name": full_name, "grade": grade, "phone": phone})
            if not valid:
                errors.append(f"Row {i + 1}: {err}")
                continue

            profile, creds = _create_student_with_credentials(
                org, full_name, grade, phone, created_by=request.user
            )
            if profile is None:
                errors.append(f"Row {i + 1}: could not generate unique email for {full_name}")
                continue

            created += 1
            credentials_list.append(creds)

    return Response({
        "created": created,
        "errors": errors[:50],
        "credentials": credentials_list,
    }, status=status.HTTP_200_OK)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsTeacher])
def bulk_import_students_view(request):
    """
    POST /api/teacher/bulk-import
    Legacy: direct import without preview. Body: multipart/form-data with file (CSV or .xlsx)
    """
    if "file" not in request.FILES and "csv" not in request.FILES:
        return Response({"detail": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)
    file = request.FILES.get("file") or request.FILES.get("csv")
    name = (file.name or "").lower()

    if name.endswith(".csv"):
        try:
            content = file.read().decode("utf-8-sig")
        except UnicodeDecodeError:
            return Response({"detail": "Invalid encoding. Use UTF-8."}, status=status.HTTP_400_BAD_REQUEST)
        rows, err = _parse_csv(content)
    elif name.endswith(".xlsx"):
        content_bytes = file.read()
        rows, err = _parse_excel(content_bytes)
    else:
        return Response({"detail": "File must be CSV or Excel (.xlsx)"}, status=status.HTTP_400_BAD_REQUEST)

    if err:
        return Response({"detail": err}, status=status.HTTP_400_BAD_REQUEST)
    if not rows:
        return Response({"detail": "File has no data rows"}, status=status.HTTP_400_BAD_REQUEST)

    org = request.user.organization
    valid_rows, _ = _get_preview(rows, org)

    created = 0
    errors = []
    credentials_list = []

    with transaction.atomic():
        for row in valid_rows:
            full_name = row["full_name"]
            grade = row.get("grade")
            phone = row.get("phone")
            profile, creds = _create_student_with_credentials(
                org, full_name, grade, phone, created_by=request.user
            )
            if profile is None:
                errors.append(f"Could not generate unique email for {full_name}")
                continue
            created += 1
            credentials_list.append(creds)

    format_param = request.query_params.get("format", "")
    if format_param.lower() == "csv" and credentials_list:
        return _credentials_csv_response(credentials_list)

    return Response({
        "created": created,
        "errors": errors[:30],
        "credentials": credentials_list,
    }, status=status.HTTP_200_OK)


def _credentials_csv_response(credentials_list):
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[
        "fullName", "studentEmail", "studentPassword", "parentEmail", "parentPassword"
    ])
    writer.writeheader()
    writer.writerows(credentials_list)
    response = HttpResponse(
        "\ufeff" + output.getvalue(),
        content_type="text/csv; charset=utf-8",
    )
    response["Content-Disposition"] = 'attachment; filename="credentials.csv"'
    return response
